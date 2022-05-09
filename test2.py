from asyncio import events
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import time
import openpyxl 
from datetime import datetime


def getExcel(book):
    result = []
    # book = openpyxl.load_workbook(path)
    sheet = book.active
    a1 = sheet['A1']
    names1 = sheet['D7':'D22']
    names2 = sheet['D24':'D39']
    names3 = sheet['D41':'D55']


    for item in names1:
        result.append({"name":item[0].value,"status":"unknown"})
    for item in names2:
        result.append({"name":item[0].value,"status":"unknown"})
    for item in names3:
        result.append({"name":item[0].value,"status":"unknown"})
    return result

def makeExcel(listOfHits,book):
    
    sheet = book.active
    a1 = sheet['A1']
    values1 = sheet['E7':'E22']
    values2 = sheet['E24':'E39']
    values3 = sheet['E41':'E55']
    i=len(values1)
    for j in range(len(values1)):
        sheet.cell(row=7+j,column = 5).value = listOfHits[j]["status"]
    for j in range(len(values2)):
        sheet.cell(row=24+j,column=5).value = listOfHits[i+j]["status"]
    i = i + len(values2)
    for j in range(len(values3)):
        sheet.cell(row=41+j,column=5).value = listOfHits[i+j]["status"]


class Hit:
    def __init__(self, timeStamp, deviceType,deviceOs,deviceApp,deviceplayer,_type,status):
        self._timeStamp = timeStamp 
        self._deviceType = deviceType
        self._deviceOs = deviceOs
        self._deviceApp = deviceApp
        self._deviceplayer = deviceplayer
        self._type = _type
        self._status = status

    def printSelf(self):
        print(self._timeStamp,self._deviceType,self._deviceOs,self._deviceApp,self._deviceplayer,self._type,self._status)

 
DRIVER_PATH = 'chromedriver'
driver = webdriver.Chrome(executable_path=DRIVER_PATH)

_id = input("Podaj numer identyfikacyjny")
# 79984241
driver.get("http://kibana.kantei-bd.redefine.pl/app/kibana#/discover?_g=(refreshInterval:(pause:!f,value:10000),time:(from:now-1h,mode:quick,to:now))&_a=(columns:!(data.userAgentData.deviceType,data.userAgentData.os,data.userAgentData.application,data.userAgentData.player,type,data.status),filters:!(('$state':(store:appState),meta:(alias:!n,disabled:!f,index:af768b90-c5ac-11ea-bebc-19fef6ef9198,key:object.id,negate:!f,params:(query:'[{}]',type:phrase),type:phrase,value:'{}'),query:(match:(object.id:(query:'{}',type:phrase))))),index:be38bb90-c5ab-11ea-bebc-19fef6ef9198,interval:auto,query:(language:lucene,query:''),sort:!(emissionDate,desc))".format(_id,_id,_id))

print("#0")
for i in range(35):
    print("time",i)
    time.sleep(1)
    try:
        print(len(driver.find_elements_by_xpath(".//td[@class='kbnDocTableCell__dataField eui-textBreakAll eui-textBreakWord']")))
    except:
        pass


times = driver.find_elements(by=By.XPATH,value = "//td[@class='eui-textNoWrap']")
values = driver.find_elements_by_xpath(".//td[@class='kbnDocTableCell__dataField eui-textBreakAll eui-textBreakWord']")
collumns = driver.find_elements_by_xpath(".//span[@i18n-id='common.ui.docTable.tableHeader.timeHeaderCellTitle']")
# print("Columns",len(collumns))
# for col in collumns:
#     print(col.text)
# print("########################")
print("times",len(times))
# for time1 in times:
#     # realTime = time1.find_element_by_xpath("//span[.='ng-non-bindable']")
#     print(time1.text)
print("########################")
print("values",len(values))
# for val in values:
#     # value = val.find_element_by_xpath("//span[.='ng-non-bindable']")
#     print(val.text)
listOfHits = []
for item in range(0,len(values),6):
    tempItem = Hit(times[int(item/6)].text,values[item+0].text,values[item+1].text,values[item+2].text,values[item+3].text,values[item+4].text,values[item+5].text)
    listOfHits.append(tempItem)

for item in listOfHits:
    item.printSelf()

book = openpyxl.load_workbook("statystyki.xlsx")
excelData = getExcel(book)
for hit in listOfHits:
    for item in excelData:
        if hit._type == item["name"]:
            item["status"] = hit._status

for item in excelData:
    print(item["name"],item["status"])

makeExcel(excelData,book)
now = datetime.now()
dt_string = now.strftime("%d_%m_%Y_%H_%M_%S")
book.save("statystykiOut"+dt_string+".xlsx")
driver.close()

print("Koniec")

