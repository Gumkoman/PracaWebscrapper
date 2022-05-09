import tkinter as tk
from tkinter import StringVar, ttk
from asyncio import events
from selenium import webdriver
# from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as ec
import time
import openpyxl
# import os
# import json
from datetime import datetime
# import pkg_resources.py2_warn

class Hit:
    def __init__(self, timeStamp, deviceType,deviceOs,deviceApp,deviceplayer,_type,status):
        self._timeStamp = timeStamp 
        self._deviceType = deviceType
        self._deviceOs = deviceOs
        self._deviceApp = deviceApp
        self._deviceplayer = deviceplayer
        self._type = _type
        self._status = status

    def printSelf(self):
        print(self._timeStamp,self._deviceType,self._deviceOs,self._deviceApp,self._deviceplayer,self._type,self._status)


class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.geometry("350x150")
        self.title('Kibana Webscrapper')
        self.resizable(0, 0)

        # configure the grid
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=2)
        self.data={}
        self.create_widgets()

    def selectChromePath(self):
        pass

    def selectOutPutPath(self):
        pass

    def getExcel(self,book):
        result = []
        # book = openpyxl.load_workbook(path)
        sheet = book.active
        a1 = sheet['A1']
        names1 = sheet['D7':'D22']
        names2 = sheet['D24':'D39']
        names3 = sheet['D41':'D55']


        for item in names1:
            result.append({"name":item[0].value,"status":"unknown"})
        for item in names2:
            result.append({"name":item[0].value,"status":"unknown"})
        for item in names3:
            result.append({"name":item[0].value,"status":"unknown"})
        return result

    def makeExcel(self,listOfHits,book):
        sheet = book.active
        a1 = sheet['A1']
        values1 = sheet['E7':'E22']
        values2 = sheet['E24':'E39']
        values3 = sheet['E41':'E55']
        i=len(values1)
        for j in range(len(values1)):
            sheet.cell(row=7+j,column = 5).value = listOfHits[j]["status"]
        for j in range(len(values2)):
            sheet.cell(row=24+j,column=5).value = listOfHits[i+j]["status"]
        i = i + len(values2)
        for j in range(len(values3)):
            sheet.cell(row=41+j,column=5).value = listOfHits[i+j]["status"]
        
        
    
    def runKibanaTest(self):
        print("kibana")
        self.outputText.set("Rozpoczynam generowanie pliku ...")
        DRIVER_PATH = 'resources/chromedriver'
        driver = webdriver.Chrome(executable_path=DRIVER_PATH)
        _id = self.IdText.get()
        print(_id)
        driver.get("http://kibana.kantei-bd.redefine.pl/app/kibana#/discover?_g=(refreshInterval:(pause:!f,value:10000),time:(from:now-1h,mode:quick,to:now))&_a=(columns:!(data.userAgentData.deviceType,data.userAgentData.os,data.userAgentData.application,data.userAgentData.player,type,data.status),filters:!(('$state':(store:appState),meta:(alias:!n,disabled:!f,index:af768b90-c5ac-11ea-bebc-19fef6ef9198,key:object.id,negate:!f,params:(query:'[{}]',type:phrase),type:phrase,value:'{}'),query:(match:(object.id:(query:'{}',type:phrase))))),index:be38bb90-c5ab-11ea-bebc-19fef6ef9198,interval:auto,query:(language:lucene,query:''),sort:!(emissionDate,desc))".format(_id,_id,_id))
        flagIsFound = False
        for i in range(35):
            print("time",i)
            time.sleep(1)
            try:
                lenTemp = len(driver.find_elements_by_xpath(".//td[@class='kbnDocTableCell__dataField eui-textBreakAll eui-textBreakWord']"))
                if lenTemp>0:
                    flagIsFound = True
                    break
            except:
                pass
        if(flagIsFound==False):
            self.outputText.set("Nie znaleziono hitów dla danego ID")
            time.sleep(5)
            driver.close()
        times = driver.find_elements(by=By.XPATH,value = "//td[@class='eui-textNoWrap']")
        values = driver.find_elements_by_xpath(".//td[@class='kbnDocTableCell__dataField eui-textBreakAll eui-textBreakWord']")
        collumns = driver.find_elements_by_xpath(".//span[@i18n-id='common.ui.docTable.tableHeader.timeHeaderCellTitle']")
        listOfHits = []
        for item in range(0,len(values),6):
            tempItem = Hit(times[int(item/6)].text,values[item+0].text,values[item+1].text,values[item+2].text,values[item+3].text,values[item+4].text,values[item+5].text)
            listOfHits.append(tempItem)

        for item in listOfHits:
            item.printSelf()

        book = openpyxl.load_workbook("resources/statystyki.xlsx")
        excelData = self.getExcel(book)
        for hit in listOfHits:
            for item in excelData:
                if hit._type == item["name"]:
                    item["status"] = hit._status

        for item in excelData:
            print(item["name"],item["status"])

        self.makeExcel(excelData,book)
        now = datetime.now()
        dt_string = now.strftime("%d_%m_%Y_%H_%M_%S")
        book.save("statystykiOut"+dt_string+".xlsx")
        driver.close()
        self.outputText.set("Plik zapisano jako:"+"statystykiOut"+dt_string+".xlsx")




    def create_widgets(self):
        # Id
        self.IdLabel = ttk.Label(self, text="ID:")
        self.IdLabel.grid(column=0, row=0, sticky=tk.W, padx=5, pady=5)

        self.IdText = ttk.Entry(self)
        self.IdText.grid(column=1, row=0, sticky=tk.E, padx=5, pady=5)
        self.IdText.insert(tk.END,"Tu wpisz numer id")
        
        self.outputText = StringVar()
        self.outputLabel = ttk.Label(self,textvariable=self.outputText)
        self.outputLabel.grid(column=0, row=2, columnspan=2,sticky=tk.E, padx=5, pady=5)
        
        # Generate
        self.generateBtn = ttk.Button(self, text="Generate", command=self.runKibanaTest)
        self.generateBtn.grid(column=0, row=1, columnspan=2,sticky="WSEN", padx=5, pady=5)

if __name__ == "__main__":
    app = App()
    app.mainloop()