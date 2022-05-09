import openpyxl 


def getExcel(book):
    result = []
    # book = openpyxl.load_workbook(path)
    sheet = book.active
    a1 = sheet['A1']
    names1 = sheet['D7':'D22']
    names2 = sheet['D24':'D39']
    names3 = sheet['D41':'D55']


    for item in names1:
        result.append({"name":item[0].value,"status":"unknown"})
    for item in names2:
        result.append({"name":item[0].value,"status":"unknown"})
    for item in names3:
        result.append({"name":item[0].value,"status":"unknown"})
    return result


def makeExcel(listOfHits,book):
    
    sheet = book.active
    a1 = sheet['A1']
    values1 = sheet['E7':'E22']
    values2 = sheet['E24':'E39']
    values3 = sheet['E41':'E55']
    i=len(values1)
    for j in range(len(values1)):
        sheet.cell(row=7+j,column = 5).value = listOfHits[j]["status"]
    for j in range(len(values2)):
        sheet.cell(row=24+j,column=5).value = listOfHits[i+j]["status"]
    i = i + len(values2)
    for j in range(len(values3)):
        sheet.cell(row=41+j,column=5).value = listOfHits[i+j]["status"]

    # for item in values1:
    #     item = listOfHits[i]['status']
    # for item in values2:
    #     item = listOfHits[i+15]['status']
    # for item in values3:
    #     item = listOfHits[i+30]['status']

book = openpyxl.load_workbook("statystyki.xlsx")

data = getExcel(book)

for i,item in enumerate(data):
    if i%2 == 0:
        item['status'] = "ok"
testLits = data

for item in testLits:
    print(item)

makeExcel(testLits,book)

book.save("statystykiOut.xlsx")

